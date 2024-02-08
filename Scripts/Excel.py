import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

#####################################################################################################################################
                                                            #Table Groupes
#####################################################################################################################################

csv_file_path_groupes = '../CSV/Groupes/Base_Groupes/Groupes.csv'

data_groupes = []
headers_groupes = []

with open(csv_file_path_groupes, newline='', encoding='utf-8') as csvfile_groupes:
    csvreader_groupes = csv.reader(csvfile_groupes, delimiter='|')
    headers_groupes = next(csvreader_groupes)
    data_groupes = list(csvreader_groupes)

wb = Workbook()

ws_groupes = wb.active
ws_groupes.title = "Groupes"

ws_groupes.append(headers_groupes)

for row_groupes in data_groupes:
    ws_groupes.append(row_groupes)

column_widths_groupes = {'A': 10, 'B': 20, 'C': 15, 'D': 55, 'E': 280, 'F': 550}
for col_groupes, width_groupes in column_widths_groupes.items():
    ws_groupes.column_dimensions[col_groupes].width = width_groupes

tab_groupes = Table(displayName="Veille", ref=f"A1:{get_column_letter(len(headers_groupes))}{len(data_groupes) + 1}")

style_groupes = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_groupes.tableStyleInfo = style_groupes

ws_groupes.add_table(tab_groupes)

#####################################################################################################################################
                                                            #Table Attaques
#####################################################################################################################################

csv_file_path_attaques = '../CSV/Attaques/Base_Attaques/Attaques.csv'

data_attaques = []
headers_attaques = []

ws_attaques = wb.create_sheet(title="Attaques Récentes")

with open(csv_file_path_attaques, newline='', encoding='utf-8') as csvfile_attaques:
    csvreader_attaques = csv.reader(csvfile_attaques, delimiter='|')
    headers_attaques = next(csvreader_attaques)
    data_attaques = list(csvreader_attaques)

ws_attaques.append(headers_attaques)

for row_attaques in data_attaques:
    ws_attaques.append(row_attaques)

column_widths_attaques = {'A': 10, 'B': 30, 'C': 15, 'D': 35, 'E': 35, 'F': 15, 'G': 550, 'H': 50, 'I': 120}
for col_attaques, width_attaques in column_widths_attaques.items():
    ws_attaques.column_dimensions[col_attaques].width = width_attaques

tab_attaques = Table(displayName="Attaques", ref=f"A1:{get_column_letter(len(headers_attaques))}{len(data_attaques) + 1}")

style_attaques = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_attaques.tableStyleInfo = style_attaques

ws_attaques.add_table(tab_attaques)

#####################################################################################################################################
                                                            #Table Victimes
#####################################################################################################################################

csv_file_path_victimes = '../CSV/Victimes/Ransomwarelive.csv'

data_victimes = []
headers_victimes = []

ws_victimes = wb.create_sheet(title="Victimes")

with open(csv_file_path_victimes, newline='', encoding='utf-8') as csvfile_victimes:
    csvreader_victimes = csv.reader(csvfile_victimes, delimiter='|')
    headers_victimes = next(csvreader_victimes)
    data_victimes = list(csvreader_victimes)

ws_victimes.append(headers_victimes)

for row_victimes in data_victimes:
    ws_victimes.append(row_victimes)

column_widths_victimes = {'A': 10, 'B': 60, 'C': 30, 'D': 15, 'E': 10, 'F': 140, 'G': 550, 'H': 150}
for col_victimes, width_victimes in column_widths_victimes.items():
    ws_victimes.column_dimensions[col_victimes].width = width_victimes

tab_victimes = Table(displayName="Victimes", ref=f"A1:{get_column_letter(len(headers_victimes))}{len(data_victimes) + 1}")

style_victimes = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_victimes.tableStyleInfo = style_victimes

ws_victimes.add_table(tab_victimes)

wb.save("../Excel/Veille.xlsx")