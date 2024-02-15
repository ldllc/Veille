import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

#####################################################################################################################################
                                                            #Table Groupes
#####################################################################################################################################

csv_file_path_groupes = '../../CSV/Groupes/Base_Groupes/Groupes.csv'

data_groupes = []
headers_groupes = []

with open(csv_file_path_groupes, newline='', encoding='utf-8') as csvfile_groupes:
    csvreader_groupes = csv.reader(csvfile_groupes, delimiter='|')
    headers_groupes = next(csvreader_groupes)
    data_groupes = list(csvreader_groupes)

wb = Workbook()

ws_groupes = wb.active
ws_groupes.title = "Groupes"

ws_groupes.append(headers_groupes)

for row_groupes in data_groupes:
    ws_groupes.append(row_groupes)

column_widths_groupes = {'A': 10, 'B': 20, 'C': 15, 'D': 55, 'E': 280, 'F': 550}
for col_groupes, width_groupes in column_widths_groupes.items():
    ws_groupes.column_dimensions[col_groupes].width = width_groupes

tab_groupes = Table(displayName="Veille", ref=f"A1:{get_column_letter(len(headers_groupes))}{len(data_groupes) + 1}")

style_groupes = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_groupes.tableStyleInfo = style_groupes

ws_groupes.add_table(tab_groupes)

#####################################################################################################################################
                                                            #Table Attaques
#####################################################################################################################################

csv_file_path_attaques = '../../CSV/Attaques/Base_Attaques/Attaques.csv'

data_attaques = []
headers_attaques = []

ws_attaques = wb.create_sheet(title="Attaques Récentes")

with open(csv_file_path_attaques, newline='', encoding='utf-8') as csvfile_attaques:
    csvreader_attaques = csv.reader(csvfile_attaques, delimiter='|')
    headers_attaques = next(csvreader_attaques)
    data_attaques = list(csvreader_attaques)

ws_attaques.append(headers_attaques)

for row_attaques in data_attaques:
    ws_attaques.append(row_attaques)

column_widths_attaques = {'A': 10, 'B': 30, 'C': 15, 'D': 35, 'E': 35, 'F': 15, 'G': 550, 'H': 50, 'I': 120}
for col_attaques, width_attaques in column_widths_attaques.items():
    ws_attaques.column_dimensions[col_attaques].width = width_attaques

tab_attaques = Table(displayName="Attaques", ref=f"A1:{get_column_letter(len(headers_attaques))}{len(data_attaques) + 1}")

style_attaques = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_attaques.tableStyleInfo = style_attaques

ws_attaques.add_table(tab_attaques)

#####################################################################################################################################
                                                            #Table Victimes
#####################################################################################################################################

csv_file_path_victimes = '../../CSV/Stats/Victimes/Victimes_Retail.csv'

data_victimes = []
headers_victimes = []

ws_victimes = wb.create_sheet(title="Victimes Retail")

with open(csv_file_path_victimes, newline='', encoding='utf-8') as csvfile_victimes:
    csvreader_victimes = csv.reader(csvfile_victimes, delimiter='|')
    headers_victimes = next(csvreader_victimes)
    data_victimes = list(csvreader_victimes)

ws_victimes.append(headers_victimes)

for row_victimes in data_victimes:
    ws_victimes.append(row_victimes)

column_widths_victimes = {'A': 10, 'B': 60, 'C': 30, 'D': 15, 'E': 10, 'F': 140, 'G': 550, 'H': 150}
for col_victimes, width_victimes in column_widths_victimes.items():
    ws_victimes.column_dimensions[col_victimes].width = width_victimes

tab_victimes = Table(displayName="Victimes", ref=f"A1:{get_column_letter(len(headers_victimes))}{len(data_victimes) + 1}")

style_victimes = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_victimes.tableStyleInfo = style_victimes

ws_victimes.add_table(tab_victimes)

#####################################################################################################################################
                                                            #Table Attaquants
#####################################################################################################################################

csv_file_path_attaquants = '../../CSV/Stats/Attaquants_Retail/Attaquants_Retail.csv'

data_attaquants = []
headers_attaquants = []
ws_attaquants = wb.create_sheet(title="Attaquants Retail")

with open(csv_file_path_attaquants, newline='', encoding='utf-8') as csvfile_attaquants:
    csvreader_attaquants = csv.reader(csvfile_attaquants, delimiter='|')
    headers_attaquants = next(csvreader_attaquants)
    data_attaquants = list(csvreader_attaquants)

ws_attaquants.append(headers_attaquants)

for row_attaquants in data_attaquants:
    ws_attaquants.append(row_attaquants)

column_widths_attaquants = {'A': 10, 'B': 10, 'C': 15, 'D': 50, 'E': 550, 'F': 15}
for col_attaquants, width_attaquants in column_widths_attaquants.items():
    ws_attaquants.column_dimensions[col_attaquants].width = width_attaquants

tab_attaquants = Table(displayName="Attaquants_Retail", ref=f"A1:{get_column_letter(len(headers_attaquants))}{len(data_attaquants) + 1}")

style_attaquants = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_attaquants.tableStyleInfo = style_attaquants

ws_attaquants.add_table(tab_attaquants)

#####################################################################################################################################
                                                            #Table Presse
#####################################################################################################################################

csv_file_path_presse = '../../CSV/Presse/Presse.csv'

data_presse = []
headers_presse = []

ws_presse = wb.create_sheet(title="Presse")

with open(csv_file_path_presse, newline='', encoding='utf-8') as csvfile_presse:
    csvreader_presse = csv.reader(csvfile_presse, delimiter='|')
    headers_presse = next(csvreader_presse)
    data_presse = list(csvreader_presse)

ws_presse.append(headers_presse)

for row_presse in data_presse:
    ws_presse.append(row_presse)

column_widths_presse = {'A': 150, 'B': 150, 'C': 15, 'D': 70}
for col_presse, width_presse in column_widths_presse.items():
    ws_presse.column_dimensions[col_presse].width = width_presse

tab_presse = Table(displayName="Presse", ref=f"A1:{get_column_letter(len(headers_presse))}{len(data_presse) + 1}")

style_presse = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_presse.tableStyleInfo = style_presse

ws_presse.add_table(tab_presse)

#####################################################################################################################################
                                                    # Table Comptes_Attaques
#####################################################################################################################################

csv_file_path_comptes_attaques = '../../CSV/Stats/Comptes_Attaques/Comptes_Attaques.csv'

data_comptes_attaques = []
headers_comptes_attaques = []
ws_comptes_attaques = wb.create_sheet(title="Comptes Attaques")

with open(csv_file_path_comptes_attaques, newline='', encoding='utf-8') as csvfile_comptes_attaques:
    csvreader_comptes_attaques = csv.reader(csvfile_comptes_attaques, delimiter='|')
    headers_comptes_attaques = next(csvreader_comptes_attaques)
    data_comptes_attaques = list(csvreader_comptes_attaques)

column_widths_compte_attaque = {'A': 30, 'B': 30}
for col_compte_attaque, width_compte_attaque in column_widths_compte_attaque.items():
    ws_comptes_attaques.column_dimensions[col_compte_attaque].width = width_compte_attaque

ws_comptes_attaques.append(headers_comptes_attaques)
for row_comptes_attaques in data_comptes_attaques:
    ws_comptes_attaques.append(row_comptes_attaques)
tab_comptes_attaques = Table(displayName="Comptes_Attaques", ref=f"A1:{get_column_letter(len(headers_comptes_attaques))}{len(data_comptes_attaques) + 1}")
style_comptes_attaques = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_comptes_attaques.tableStyleInfo = style_comptes_attaques
ws_comptes_attaques.add_table(tab_comptes_attaques)

#####################################################################################################################################
                                                # Table Nombre d'attaques par jours
#####################################################################################################################################

csv_file_path_comptes_journaliers = '../../CSV/Stats/Comptes_Journaliers/Comptes_Journaliers.csv'
ws_comptes_journaliers = wb.create_sheet(title="Comptes Journaliers")
data_comptes_journaliers = []
headers_comptes_journaliers = []

with open(csv_file_path_comptes_journaliers, newline='', encoding='utf-8') as csvfile_comptes_journaliers:
    csvreader_comptes_journaliers = csv.reader(csvfile_comptes_journaliers, delimiter='|')
    headers_comptes_journaliers = next(csvreader_comptes_journaliers)
    data_comptes_journaliers = list(csvreader_comptes_journaliers)

column_widths_compte_journalier = {'A': 30, 'B': 30}

for col_compte_journalier, width_compte_journalier in column_widths_compte_journalier.items():
    ws_comptes_journaliers.column_dimensions[col_compte_journalier].width = width_compte_journalier

ws_comptes_journaliers.append(headers_comptes_journaliers)

for row_comptes_journaliers in data_comptes_journaliers:
    ws_comptes_journaliers.append(row_comptes_journaliers)

tab_comptes_journaliers = Table(displayName="Comptes_Journaliers", ref=f"A1:{get_column_letter(len(headers_comptes_journaliers))}{len(data_comptes_journaliers) + 1}")
style_comptes_journaliers = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                          showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab_comptes_journaliers.tableStyleInfo = style_comptes_journaliers
ws_comptes_journaliers.add_table(tab_comptes_journaliers)

wb.save("../../Excel/Veille.xlsx")